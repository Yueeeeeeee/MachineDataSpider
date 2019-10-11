import os
from openpyxl import load_workbook
from xlwt import *


cwd = os.getcwd()
print(cwd)
os.chdir("/Users/z00445wp/Desktop/MachineDataSpider")
print(os.listdir("."))
excelFile = Workbook(encoding='utf-8')

excelTable = excelFile.add_sheet('MCenter')


wb = load_workbook('./Categories.xlsx')

string = ['MCenter']
MCenter_Filter = ['Rapid traverse speed X/Y/Z | Power (Speed) m/min', 'Rapid traverse speed X/Y/Z m/min',
                     "Working travels in X-/Y'-/Z-axis (mm)", "Max. speeds in X-/Y'-/Z-axis (m/min)",
                     "Positioning range X/Y/Z mm",
                     "Max. speeds (5-axis version) in X'-/Y-/Z-/W-axis (m/min)",
                     "Working travels (5-axis version) in X'-/Y-/Z-/W-axis (mm)",
                     "Max. speeds max. in X-/Y'-/Z-Achse (m/min)",
                     "Travel X, Y, Z", "Rapid traverse X, Y, Z", "Rapid traverse X, Y",
                     "Feed rate / rapid traverse X, Y, Z", "Rapide traverse Z",
                     "Feed rate / rapid traverse (swiveling, rotating)", "Feed rate / Rapid traverse Z",
                     "Max. X travels", "Max. Y travels", "Max. Z travel", "Max. X axis", "Max. Y axis", "Max. Z axis",
                     "X Axis Travel distance", "Y Axis Travel distance",	"Z Axis Travel distance",
                     "X Axis Rapid Traverse",
                     "Y Axis Rapid Traverse",	"Z Axis Rapid Traverse",
                     "X-axis travel (column right and left)",
                     "Y-axis travel (spindle up and down)",	"Z-axis travel (table back and forth)",
                     "Rapid traverse (X/Y/Z)",
                     "Maximum rapid traverse X-axis",
                     "Maximum rapid traverse Y-axis",
                     "Maximum rapid traverse Z-axis",
                     "X-axis travel (Spindle head cross-wise)",
                     "Y-axis travel (Spindle head up/down)",
                     "X-axis travel (Spindle head right/left)",
                  "Rapid traverse (swiveling, rotating)",
                  "Z-axis travel (Table back/forth)", "Z-axis travel (Spindle head back/forth)",
                  "Rapid traverse rate Z-axis", "Rapid traverse (X-axis)", "Rapid traverse (Y-axis)",
                  "Rapid traverse (Z-axis)", "Rapid traverse rate (X-axis)", "Rapid traverse rate (Y-axis)",
                  "Rapid traverse rate (Z-axis)", "Rapid traverse rate (X/Y/Z)"
                  "Rapid traverse rate X-axis",	"Rapid traverse rate Y-axis", "Rapid traverse rate Z-Axis",
                  "Rapid traverse (X)",	"Rapid traverse (Y)", "Rapid traverse (Z)",
                  "Rapid traverse rate (X-,  Y-,  Z-)", "Maximum workpiece diameter", "Maximum workpiece height",
                  "Max.material diameter", "X - axis", "Y - axis", "Maximum machining diameter",
                  "Maximum workpiece width (X)", "Maximum workpiece length (Y)",
                  "Maximum machining diameter (upper turret)", "Maximum machining length",
                  "Maximum machining diameter (lower turret)",
                  "Work table cross dimension",	"Work table longitudinal dimension", "Maximum workpiece width",
                  "Maximum workpiece length", "X axis stroke", "Y axis stroke", "Z axis stroke", "X-axis stroke",
                  "Y-axis stroke", "Z-axis stroke", "X-axis stroke (saddle right/left movement)",
                  "Y-axis stroke (column back/forth movement)", "Z-axis stroke (spindle up/down movement)",
                  "Movement stroke X", "Movement stroke Z",	"Movement stroke Y"]

MCenter_DiameterFilter = ["Working travels in X-/Y'-/Z-axis (mm)", "Positioning range X/Y/Z mm",
                         "Working travels (5-axis version) in X'-/Y-/Z-/W-axis (mm)",
                         "Travel X, Y, Z",
                         "Max. X travels", "Max. Y travels", "Max. Z travels",
                         "X Axis Travel distance", "Y Axis Travel distance",	"Z Axis Travel distance",
                         "X-axis travel (column right and left)",
                         "Y-axis travel (spindle up and down)",	"Z-axis travel (table back and forth)",
                         "X-axis travel (Spindle head cross-wise)",
                         "Y-axis travel (Spindle head up/down)",
                         "X-axis travel (Spindle head right/left)",
                      "Z-axis travel (Table back/forth)", "Z-axis travel (Spindle head back/forth)",
                      "Maximum workpiece diameter", "Maximum workpiece height",
                      "Max.material diameter", "X - axis", "Y - axis", "Maximum machining diameter",
                      "Maximum workpiece width (X)", "Maximum workpiece length (Y)",
                      "Maximum machining diameter (upper turret)", "Maximum machining length",
                      "Maximum machining diameter (lower turret)",
                      "Work table cross dimension",	"Work table longitudinal dimension", "Maximum workpiece width",
                      "Maximum workpiece length", "X axis stroke", "Y axis stroke", "Z axis stroke", "X-axis stroke",
                      "Y-axis stroke", "Z-axis stroke", "X-axis stroke (saddle right/left movement)",
                      "Y-axis stroke (column back/forth movement)", "Z-axis stroke (spindle up/down movement)",
                      "Movement stroke X", "Movement stroke Z",	"Movement stroke Y", "Z2 axis stroke", "W axis stroke",
                      "W axis (boring spindle stroke)", "A-axis travel (table tiliting)", "Movement stroke X2",
                      "Movement stroke Z2", "Movement stroke B", "X-axis stroke (Table back/forth)",
                      "Y-axis stroke (Spindle head right/left)", "Z-axis stroke (Spindle head up/down)",
                      "A-axis travel", "C-axis travel (table rotating)", "A-axis (table tilt) travel amount/indexing 0.0001°",
                      "B axis rotational stroke","C-axis (table rotation) travel amount/indexing 0.0001°",
                          "C-axis travel (standard)", "C-axis travel (optional)"]


MCenter_MaxXFilter = []
for q in range(len(MCenter_DiameterFilter)):
    if "X" in MCenter_DiameterFilter[q] and "X2" not in MCenter_DiameterFilter[q] and "(X)" not in MCenter_DiameterFilter[q]:
        MCenter_MaxXFilter.append(MCenter_DiameterFilter[q])

print(MCenter_MaxXFilter)

MCenter_MaxX2Filter = []
for q in range(len(MCenter_DiameterFilter)):
    if "X2" in MCenter_DiameterFilter[q]:
        MCenter_MaxX2Filter.append(MCenter_DiameterFilter[q])

print(MCenter_MaxX2Filter)

MCenter_MaxYFilter = []
for q in range(len(MCenter_DiameterFilter)):
    if "Y" in MCenter_DiameterFilter[q] and "(Y)" not in MCenter_DiameterFilter[q]:
        MCenter_MaxYFilter.append(MCenter_DiameterFilter[q])

print(MCenter_MaxYFilter)

MCenter_MaxZFilter = []
for q in range(len(MCenter_DiameterFilter)):
    if "Z" in MCenter_DiameterFilter[q]:
        MCenter_MaxZFilter.append(MCenter_DiameterFilter[q])

print(MCenter_MaxZFilter)

MCenter_MaxWFilter = []
for q in range(len(MCenter_DiameterFilter)):
    if "W axis" in MCenter_DiameterFilter[q] or "W-axis" in MCenter_DiameterFilter[q]:
        MCenter_MaxWFilter.append(MCenter_DiameterFilter[q])

print(MCenter_MaxWFilter)

MCenter_MaxAFilter = []
for q in range(len(MCenter_DiameterFilter)):
    if 'A-axis' in MCenter_DiameterFilter[q]:
        MCenter_MaxAFilter.append(MCenter_DiameterFilter[q])

print(MCenter_MaxAFilter)

MCenter_MaxBFilter = []
for q in range(len(MCenter_DiameterFilter)):
    if "B" in MCenter_DiameterFilter[q]:
        MCenter_MaxBFilter.append(MCenter_DiameterFilter[q])

print(MCenter_MaxBFilter)

MCenter_MaxCFilter = []
for q in range(len(MCenter_DiameterFilter)):
    if "C" in MCenter_DiameterFilter[q]:
        MCenter_MaxCFilter.append(MCenter_DiameterFilter[q])

print(MCenter_MaxCFilter)


data_row = 0
data_row_Table = 0
stop_point = 0

######### First two columns for X Max???
sheet = wb.get_sheet_by_name(string[0])
for i in range(sheet.max_row):                        # screening from top to bottom
    stop_point = 0
    data_column = 0
    for j in range(0, sheet.max_column):
        if j % 2 == 1 and stop_point == 0:
            Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
            Data = sheet.cell(row=i + 1, column=j + 2).value
            if Spec in MCenter_MaxXFilter:
                if "/Z" not in Spec and ", Z" not in Spec:
                    excelTable.write(data_row_Table, data_column, str(Spec))
                    excelTable.write(data_row_Table, data_column + 1, str(Data))
                    stop_point = 1
                elif "/Z" in Spec:
                    Spec = str(Spec).split('/')[0]
                    try:
                        Data = str(Data).split('/')[0]
                    except:
                        pass
                    excelTable.write(data_row_Table, data_column, str(Spec))
                    excelTable.write(data_row_Table, data_column + 1, str(Data))
                    stop_point = 1
                elif ", Z" in Spec:
                    Spec = str(Spec).split(',')[0]
                    try:
                        Data = str(Data).split('x')[0]
                    except:
                        pass
                    excelTable.write(data_row_Table, data_column, str(Spec))
                    excelTable.write(data_row_Table, data_column + 1, str(Data))
                    stop_point = 1

            else:
                pass
        elif j == 0 and stop_point == 0:
            Name = sheet.cell(row=i + 1, column=j + 1).value
            excelTable.write(data_row_Table, data_column, str(Name))
            data_column += 1
        else:
            pass
    data_row_Table += 1

data_row = 0
data_row_Table = 0
stop_point = 0
######### First two columns for Y Max???
for i in range(sheet.max_row):                        # screening from top to bottom
    stop_point = 0
    data_column = 3
    for j in range(0, sheet.max_column):
        if j % 2 == 1 and stop_point == 0:
            Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
            Data = sheet.cell(row=i + 1, column=j + 2).value
            if Spec in MCenter_MaxYFilter:
                if "/Z" not in Spec and ", Z" not in Spec:
                    excelTable.write(data_row_Table, data_column, str(Spec))
                    excelTable.write(data_row_Table, data_column + 1, str(Data))
                    stop_point = 1
                elif "/Z" in Spec:
                    Spec = str(Spec).split('/')[1]
                    try:
                        Data = str(Data).split('/')[1]
                    except:
                        pass
                    excelTable.write(data_row_Table, data_column, str(Spec))
                    excelTable.write(data_row_Table, data_column + 1, str(Data))
                    stop_point = 1
                elif ", Z" in Spec:
                    Spec = str(Spec).split(',')[1]
                    try:
                        Data = str(Data).split('x')[1]
                    except:
                        pass
                    excelTable.write(data_row_Table, data_column, str(Spec))
                    excelTable.write(data_row_Table, data_column + 1, str(Data))
                    stop_point = 1
            else:
                pass
        else:
            pass
    data_row_Table += 1

data_row = 0
data_row_Table = 0
stop_point = 0
######### First two columns for Z Max???
for i in range(sheet.max_row):                        # screening from top to bottom
    stop_point = 0
    data_column = 5
    for j in range(0, sheet.max_column):
        if j % 2 == 1 and stop_point == 0:
            Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
            Data = sheet.cell(row=i + 1, column=j + 2).value
            if Spec in MCenter_MaxZFilter:
                if "/Z" not in Spec and ", Z" not in Spec:
                    excelTable.write(data_row_Table, data_column, str(Spec))
                    excelTable.write(data_row_Table, data_column + 1, str(Data))
                    stop_point = 1
                elif "/Z" in Spec:
                    Spec = str(Spec).split('/')[2]
                    try:
                        Data = str(Data).split('/')[2]
                    except:
                        pass
                    excelTable.write(data_row_Table, data_column, str(Spec))
                    excelTable.write(data_row_Table, data_column + 1, str(Data))
                    stop_point = 1
                elif ", Z" in Spec:
                    Spec = str(Spec).split(',')[2]
                    try:
                        Data = str(Data).split('x')[2]
                    except:
                        pass
                    excelTable.write(data_row_Table, data_column, str(Spec))
                    excelTable.write(data_row_Table, data_column + 1, str(Data))
                    stop_point = 1
            else:
                pass
        else:
            pass
    data_row_Table += 1

data_row = 0
data_row_Table = 0
stop_point = 0
######### First two columns for W Max???
for i in range(sheet.max_row):                        # screening from top to bottom
    stop_point = 0
    data_column = 7
    for j in range(0, sheet.max_column):
        if j % 2 == 1 and stop_point == 0:
            Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
            Data = sheet.cell(row=i + 1, column=j + 2).value
            if Spec in MCenter_MaxWFilter:
                if "/Z" not in Spec and ", Z" not in Spec:
                    excelTable.write(data_row_Table, data_column, str(Spec))
                    excelTable.write(data_row_Table, data_column + 1, str(Data))
                    stop_point = 1
                elif "/Z" in Spec:
                    Spec = str(Spec).split('/')[3]
                    try:
                        Data = str(Data).split('/')[3]
                    except:
                        pass
                    excelTable.write(data_row_Table, data_column, str(Spec))
                    excelTable.write(data_row_Table, data_column + 1, str(Data))
                    stop_point = 1
            else:
                pass
        else:
            pass
    data_row_Table += 1

data_row = 0
data_row_Table = 0
stop_point = 0
######### First two columns for A Max???
for i in range(sheet.max_row):                        # screening from top to bottom
    stop_point = 0
    data_column = 9
    for j in range(0, sheet.max_column):
        if j % 2 == 1 and stop_point == 0:
            Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
            Data = sheet.cell(row=i + 1, column=j + 2).value
            if Spec in MCenter_MaxAFilter:
                excelTable.write(data_row_Table, data_column, str(Spec))
                excelTable.write(data_row_Table, data_column + 1, str(Data))
                stop_point = 1
            else:
                pass
        else:
            pass
    data_row_Table += 1

data_row = 0
data_row_Table = 0
stop_point = 0
######### First two columns for B Max???
for i in range(sheet.max_row):                        # screening from top to bottom
    stop_point = 0
    data_column = 11
    for j in range(0, sheet.max_column):
        if j % 2 == 1 and stop_point == 0:
            Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
            Data = sheet.cell(row=i + 1, column=j + 2).value
            if Spec in MCenter_MaxBFilter:
                excelTable.write(data_row_Table, data_column, str(Spec))
                excelTable.write(data_row_Table, data_column + 1, str(Data))
                stop_point = 1
            else:
                pass
        else:
            pass
    data_row_Table += 1


data_row = 0
data_row_Table = 0
stop_point = 0
######### First two columns for C Max???
for i in range(sheet.max_row):                        # screening from top to bottom
    stop_point = 0
    data_column = 13
    for j in range(0, sheet.max_column):
        if j % 2 == 1 and stop_point == 0:
            Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
            Data = sheet.cell(row=i + 1, column=j + 2).value
            if Spec in MCenter_MaxCFilter:
                excelTable.write(data_row_Table, data_column, str(Spec))
                excelTable.write(data_row_Table, data_column + 1, str(Data))
                stop_point = 1
            else:
                pass
        else:
            pass
    data_row_Table += 1
# for q in range(len(string)
# for q in range(len(string)):
#     sheet = wb.get_sheet_by_name(string[q])
#     for i in range(sheet.max_row):                        # screening from top to bottom
#         data_column = 0
#         for j in range(0, sheet.max_column):
#             if j % 2 == 1:
#                 Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
#                 Data = sheet.cell(row=i + 1, column=j + 2).value
#                 if Spec in MCenter_MaxXFilter:
#                     excelTable.write(data_row_Table, data_column, str(Spec))
#                     excelTable.write(data_row_Table, data_column + 1, str(Data))
#                     data_column += 2
#                 else:
#                     pass
#             elif j == 0:
#                 Name = sheet.cell(row=i + 1, column=j + 1).value
#                 excelTable.write(data_row_Table, data_column, str(Name))
#                 data_column += 1
#             else:
#                 pass
#
#         data_row_Table += 1





excelFile.save('Filter_2.0.xls')

