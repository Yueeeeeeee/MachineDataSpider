import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from xlwt import *
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import YELLOW

cwd = os.getcwd()
print(cwd)
os.chdir("/Users/qinhaochen/Documents")
print(os.listdir("."))
excelFile = Workbook(encoding='utf-8')
excelTable = excelFile.add_sheet('Specifications')

wb = load_workbook('./Alles_in_Allem.xlsx')

string = ['Mazak', 'Index-Werke', 'Heller', 'Doosan', 'GROB', 'DMG', 'G+F']
data_row = 0

for q in range(len(string)):
    data_column = 0
    sheet = wb.get_sheet_by_name(string[q])
    Label_List = []
    for i in range(sheet.max_row):
        for j in range(sheet.max_column):
            if j % 2 == 1:
                Spec = sheet.cell(row=i+1, column=j+1).value
                if str(Spec) not in Label_List:
                    Label_List.append(str(Spec))
                else:
                    pass
            else:
                pass
    excelTable.write(data_row, 0, string[q])
    sheet.cell(row=data_row+1, column=1).fill = PatternFill(fgColor=YELLOW, fill_type = "solid")
    data_row += 1
    for r in range(len(Label_List)):
        excelTable.write(data_row, data_column, Label_List[r])
        data_column += 1
        if data_column == 5:
            data_column = 0
            data_row += 1
        else:
            pass
    data_row += 1

excelFile.save('Specifications.xls')
