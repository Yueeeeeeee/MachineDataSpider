
from openpyxl import load_workbook
from xlwt import *
import os



cwd = os.getcwd()
print(cwd)
os.chdir("/Users/z00445wp/Desktop/PPT/06122019")
print(os.listdir("."))
excelFile = Workbook(encoding='utf-8')
excelTable = excelFile.add_sheet('Summe')

wb = load_workbook('./PVO SPLIT PER SUBCOMMODITY_FY19.xlsx')

string = ['5.Project Material', '4. Additive', '3. Robotic', '2. Mechanic Product.', '1. Electronic Product.', 'All_PVO']

data_row = 0

for i in range(len(string)):
    sheet = wb.get_sheet_by_name(string[i])
    excelTable.write(data_row, 0, str(string[i]))
    Sum = 0
    for j in range(2, sheet.max_row+1):
        if sheet.cell(row=j, column=5).value == "RMB":
            RMB = sheet.cell(row=j, column=11).value
            Euro = RMB / 7.88
            Sum += Euro
        elif sheet.cell(row=j, column=5).value == "USD":
            USD = sheet.cell(row=j, column=11).value
            Euro = USD / 1.11
            Sum += Euro
        elif sheet.cell(row=j, column=5).value == "GBP":
            GBP = sheet.cell(row=j, column=11).value
            Euro = GBP * 1.19
            Sum += Euro
        elif sheet.cell(row=j, column=5).value == "DKK":
            DKK = sheet.cell(row=j, column=11).value
            Euro = DKK * 0.13
            Sum += Euro
        elif sheet.cell(row=j, column=5).value is None:
            Euro = 0
            Sum += 0
        else:
            Euro = sheet.cell(row=j, column=11).value
            Sum += Euro
    excelTable.write(data_row, 1, Sum)
    data_row += 1

excelFile.save('Sum.xls')
