import os
from openpyxl import load_workbook
from xlwt import *


cwd = os.getcwd()
print(cwd)
os.chdir("/Users/z00445wp/Desktop/MachineDataSpider")
print(os.listdir("."))
excelFile = Workbook(encoding='utf-8')
excelTable = excelFile.add_sheet('Specifications')

wb = load_workbook('./Alles_in_Allem.xlsx')

string = ['Mazak', 'Index-Werke', 'Heller', 'Doosan', 'GROB', 'DMG', 'G+F']
data_row = 0

for q in range(len(string)):
    data_column = 0
    sheet = wb.get_sheet_by_name(string[q])
    Label_List_Cat1 = []
    Label_List_Cat2 = []
    Label_List_Cat3 = []

    for i in range(sheet.max_row):                        # screening from top to bottom
        for j in range(1, sheet.max_column):              # screening from left to right
            if sheet.cell(row=i+1, column=1).value == 1:   # Identifying the machine type
                if j % 2 == 0:                                 #Screening the specs, save, unless it shows before
                    Spec = sheet.cell(row=i+1, column=j+1).value
                    if str(Spec) not in Label_List_Cat1 and Spec is not None:
                        Label_List_Cat1.append(str(Spec))
                    else:
                        pass
                else:
                    pass
            elif sheet.cell(row=i+1, column=1).value == 2:
                if j % 2 == 0:
                    Spec = sheet.cell(row=i+1, column=j+1).value
                    if str(Spec) not in Label_List_Cat2 and Spec is not None:
                        Label_List_Cat2.append(str(Spec))
                    else:
                        pass
                else:
                    pass
            elif sheet.cell(row=i+1, column=1).value == 3:
                if j % 2 == 0:
                    Spec = sheet.cell(row=i+1, column=j+1).value
                    if str(Spec) not in Label_List_Cat3 and Spec is not None:
                        Label_List_Cat3.append(str(Spec))
                    else:
                        pass
                else:
                    pass

    excelTable.write(data_row, 0, string[q])
    data_row += 1
    excelTable.write(data_row, 0, 1)
    data_row += 1
    for r in range(len(Label_List_Cat1)):
        excelTable.write(data_row, data_column, Label_List_Cat1[r])
        data_column += 1
        if data_column == 4:
            data_column = 0
            data_row += 1
        else:
            pass
    data_row += 1
    excelTable.write(data_row, 0, 2)
    data_row += 1
    data_column = 0
    for x in range(len(Label_List_Cat2)):
        excelTable.write(data_row, data_column, Label_List_Cat2[x])
        data_column += 1
        if data_column == 4:
            data_column = 0
            data_row += 1
        else:
            pass
    data_row += 1
    excelTable.write(data_row, 0, 3)
    data_row += 1
    data_column = 0
    for y in range(len(Label_List_Cat3)):
        excelTable.write(data_row, data_column, Label_List_Cat3[y])
        data_column += 1
        if data_column == 4:
            data_column = 0
            data_row += 1
        else:
            pass
    data_row += 1

excelFile.save('Specifications.xls')
