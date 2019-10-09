import os
from openpyxl import load_workbook
from xlwt import *


cwd = os.getcwd()
print(cwd)
os.chdir("/Users/z00445wp/Desktop/MachineDataSpider")
print(os.listdir("."))
excelFile = Workbook(encoding='utf-8')
excelTable1 = excelFile.add_sheet('Turning')
excelTable2 = excelFile.add_sheet('MCenter')
excelTable3 = excelFile.add_sheet('Combi')

wb = load_workbook('./Alles_in_Allem.xlsx')

string = ['Mazak', 'Index-Werke', 'Heller', 'Doosan', 'GROB', 'DMG', 'G+F']
# data_row = 0
data_row_Table1 = 0
data_row_Table2 = 0
data_row_Table3 = 0

for q in range(len(string)):
    sheet = wb.get_sheet_by_name(string[q])
    for i in range(sheet.max_row):                        # screening from top to bottom
        data_column = 0

        if sheet.cell(row=i + 1, column=1).value == 1:  # Identifying the machine type
            for j in range(1, sheet.max_column):
                if j == 1:
                    Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
                    excelTable1.write(data_row_Table1, data_column, string[q])
                    excelTable1.write(data_row_Table1, data_column + 1, str(Spec))
                    data_column += 2
                else:
                    Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
                    if Spec is None and sheet.cell(row=i + 1, column=j + 2).value is None:
                        break
                    else:
                        excelTable1.write(data_row_Table1, data_column, str(Spec))
                        data_column += 1
            data_row_Table1 += 1

        elif sheet.cell(row=i + 1, column=1).value == 2:  # Identifying the machine type
            for j in range(1, sheet.max_column):
                if j == 1:
                    Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
                    excelTable2.write(data_row_Table2, data_column, string[q])
                    excelTable2.write(data_row_Table2, data_column + 1, str(Spec))
                    data_column += 2
                else:
                    Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
                    if Spec is None and sheet.cell(row=i + 1, column=j + 2).value is None:
                        break
                    else:
                        excelTable2.write(data_row_Table2, data_column, str(Spec))
                        data_column += 1
            data_row_Table2 += 1

        elif sheet.cell(row=i + 1, column=1).value == 3:  # Identifying the machine type
            for j in range(1, sheet.max_column):
                if j == 1:
                    Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
                    excelTable3.write(data_row_Table3, data_column, string[q])
                    excelTable3.write(data_row_Table3, data_column + 1, str(Spec))
                    data_column += 2
                else:
                    Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
                    if Spec is None and sheet.cell(row=i + 1, column=j + 2).value is None:
                        break
                    else:
                        excelTable3.write(data_row_Table3, data_column, str(Spec))
                        data_column += 1
            data_row_Table3 += 1

excelFile.save('Categories.xls')
