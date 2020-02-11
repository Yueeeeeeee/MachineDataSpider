import os
import re
from openpyxl import load_workbook
from xlwt import *


cwd = os.getcwd()
print(cwd)
os.chdir("/Users/z00445wp/Desktop/Report/Lieferanten201819")
print(os.listdir("."))
excelFile = Workbook(encoding='utf-8')
excelTable = excelFile.add_sheet('Top_15')

wb = load_workbook('./1819_Liefranten_PEPDPMF.xlsx')

string = ['Raw_Data']
data_row = 0


data_column = 0
sheet = wb.get_sheet_by_name('Raw_Data')
Supplier_List = []
for i in range(1, sheet.max_row):
    Supplier = sheet.cell(row=i + 1, column=8).value
    if any(s in Supplier for s in Supplier_List if Supplier) or Supplier is None:
        pass
    else:
        Supplier_List.append(Supplier)
else:
    pass

print(Supplier_List)

for k in range(len(Supplier_List)):
    excelTable.write(data_row, 0, Supplier_List[k])
    Sum = 0
    for i in range(1, sheet.max_row):
        Supplier = sheet.cell(row=i + 1, column=8).value
        if Supplier is not None:
            if Supplier_List[k] in Supplier:
                if sheet.cell(row=i + 1, column=5).value == "RMB":
                    RMB = sheet.cell(row=i + 1, column=11).value
                    Euro = RMB/7.88
                    Sum += Euro
                elif sheet.cell(row=i + 1, column=5).value == "USD":
                    USD = sheet.cell(row=i + 1, column=11).value
                    Euro = USD/1.11
                    Sum += Euro
                elif sheet.cell(row=i + 1, column=5).value == "GBP":
                    GBP = sheet.cell(row=i + 1, column=11).value
                    Euro = GBP*1.19
                    Sum += Euro
                else:
                    if sheet.cell(row=i + 1, column=11).value is not None:
                        Euro = sheet.cell(row=i + 1, column=11).value
                        Sum += Euro
                    else:
                        pass
        else:
            pass

    excelTable.write(data_row, 1, Sum)
    data_row += 1



 # elif j ==7:                                       %Supplier
 #                Spec = sheet.cell(row=i + 1, column=j + 1).value
 #                if str(Spec) not in Label_List and Spec is not None:
 #                    Label_List.append(str(Spec))
 #                else:
 #                    pass
 # #
 #            if j == 4:                                      # Currency
 #                if sheet.cell(row=i + 1, column=j + 1).value is "RMB":
 #                    RMB = sheet.cell(row=i + 1, column=11).value
 #                    Euro = RMB/8
 #                    excelTable.write(data_row, data_column, Euro)
 #                else:
 #                    pass
 #            else:
 #                pass

excelFile.save('1819_Liefranten_PEPDPMF.xls')