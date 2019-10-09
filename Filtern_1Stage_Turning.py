import os
from openpyxl import load_workbook
from xlwt import *


cwd = os.getcwd()
print(cwd)
os.chdir("/Users/z00445wp/Desktop/MachineDataSpider")
print(os.listdir("."))
excelFile = Workbook(encoding='utf-8')

excelTable = excelFile.add_sheet('MCenter')


wb = load_workbook('./Categories.xlsx')

string = ['Turning']
Turning_Chuck = ["Chuck size", "Chuck size main spindle", "Chuck diameter  mm", "Max. chuck size"]
Turning_Number = ["Number of tools", "Number of stations -", "No. of tool station", "Max. number of tool carriers"]
Turning_Speed = ["Rotating speed maximum", "Maximum milling spindle speed", "Speed max. rpm", "Max. Spindle Speed", "Max spindle motor speed", "Max. speed"]
Turning_Power = ["Power at 100% / 40% kW", "Power max. kW"]

#


data_row = 0
data_row_Table = 0
stop_point = 0

######### First two columns for Chuck???
sheet = wb.get_sheet_by_name(string[0])
for i in range(sheet.max_row):                        # screening from top to bottom
    stop_point = 0
    data_column = 0
    for j in range(0, sheet.max_column):
        if j % 2 == 1 and stop_point == 0:
            Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
            Data = sheet.cell(row=i + 1, column=j + 2).value
            if Spec in MCenter_Chuck:
                excelTable.write(data_row_Table, data_column, str(Spec))
                excelTable.write(data_row_Table, data_column + 1, str(Data))
                stop_point = 1
            else:
                pass
        elif j == 0 and stop_point == 0:
            Name = sheet.cell(row=i + 1, column=j + 1).value
            excelTable.write(data_row_Table, data_column, str(Name))
            data_column += 1
        else:
            pass
    data_row_Table += 1

data_row = 0
data_row_Table = 0
stop_point = 0
######### First two columns for Number???
for i in range(sheet.max_row):                        # screening from top to bottom
    stop_point = 0
    data_column = 3
    for j in range(0, sheet.max_column):
        if j % 2 == 1 and stop_point == 0:
            Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
            Data = sheet.cell(row=i + 1, column=j + 2).value
            if Spec in MCenter_Number:
                excelTable.write(data_row_Table, data_column, str(Spec))
                excelTable.write(data_row_Table, data_column + 1, str(Data))
                stop_point = 1
            else:
                pass
        else:
            pass
    data_row_Table += 1

data_row = 0
data_row_Table = 0
stop_point = 0
######### First two columns for Speed???
for i in range(sheet.max_row):                        # screening from top to bottom
    stop_point = 0
    data_column = 5
    for j in range(0, sheet.max_column):
        if j % 2 == 1 and stop_point == 0:
            Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
            Data = sheet.cell(row=i + 1, column=j + 2).value
            if Spec in MCenter_Speed:
                excelTable.write(data_row_Table, data_column, str(Spec))
                excelTable.write(data_row_Table, data_column + 1, str(Data))
                stop_point = 1
            else:
                pass
        else:
            pass
    data_row_Table += 1

data_row = 0
data_row_Table = 0
stop_point = 0
######### First two columns for Power???
for i in range(sheet.max_row):                        # screening from top to bottom
    stop_point = 0
    data_column = 7
    for j in range(0, sheet.max_column):
        if j % 2 == 1 and stop_point == 0:
            Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
            Data = sheet.cell(row=i + 1, column=j + 2).value
            if Spec in MCenter_Power:
                excelTable.write(data_row_Table, data_column, str(Spec))
                excelTable.write(data_row_Table, data_column + 1, str(Data))
                stop_point = 1
            else:
                pass
        else:
            pass
    data_row_Table += 1


# for q in range(len(string)):
#     sheet = wb.get_sheet_by_name(string[q])
#     for i in range(sheet.max_row):                        # screening from top to bottom
#         data_column = 0
#         for j in range(0, sheet.max_column):
#             if j % 2 == 1:
#                 Spec = sheet.cell(row=i + 1, column=j + 1).value  # Screening the specs, save
#                 Data = sheet.cell(row=i + 1, column=j + 2).value
#                 if Spec in MCenter_MaxXFilter:
#                     excelTable.write(data_row_Table, data_column, str(Spec))
#                     excelTable.write(data_row_Table, data_column + 1, str(Data))
#                     data_column += 2
#                 else:
#                     pass
#             elif j == 0:
#                 Name = sheet.cell(row=i + 1, column=j + 1).value
#                 excelTable.write(data_row_Table, data_column, str(Name))
#                 data_column += 1
#             else:
#                 pass
#
#         data_row_Table += 1





excelFile.save('Filter_2.0_Turning.xls')

